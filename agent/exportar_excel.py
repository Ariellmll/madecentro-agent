# agent/exportar_excel.py — Generación de Excel descargable para la orden de corte
import os
import re
import logging
from openpyxl import Workbook
from openpyxl.styles import Font

logger = logging.getLogger("agentkit")

_PATRON_MATERIAL = re.compile(r'Material:\s*(.+)')
_PATRON_TOTAL_PIEZAS = re.compile(r'Total de piezas:\s*(.+)')
_PATRON_ESPESOR = re.compile(r'Espesor:\s*(.+)')

_ANCHO_COLUMNA = 18


def _parsear_tabla(cuerpo_orden: str) -> list[list[str]]:
    """Extrae filas de la tabla markdown (header + datos), sin la fila separadora."""
    filas_crudas = [
        linea.strip() for linea in cuerpo_orden.splitlines()
        if linea.strip().startswith('|')
    ]

    if not filas_crudas:
        logger.warning("No se encontró tabla markdown en el cuerpo de la orden")
        return []

    filas = []
    for linea in filas_crudas:
        # Fila separadora: solo guiones, pipes, espacios y dos puntos (ej. |---|:---:|)
        if re.fullmatch(r'[|\-:\s]+', linea):
            continue
        celdas = [c.strip() for c in linea.split('|')]
        celdas = celdas[1:-1]  # descartar los vacíos del split por los pipes de los extremos
        filas.append(celdas)

    return filas


def generar_excel_orden(cuerpo_orden: str, numero_orden: str) -> str:
    """Genera un Excel descargable con la tabla de la orden y lo guarda en generated/."""
    filas = _parsear_tabla(cuerpo_orden)

    match_material = _PATRON_MATERIAL.search(cuerpo_orden)
    match_total = _PATRON_TOTAL_PIEZAS.search(cuerpo_orden)
    match_espesor = _PATRON_ESPESOR.search(cuerpo_orden)

    wb = Workbook()
    ws = wb.active
    ws.title = "Orden de corte"

    fila_actual = 1

    if match_material:
        ws.cell(row=fila_actual, column=1, value=match_material.group(0).strip())
        fila_actual += 1

    if filas:
        header, *datos = filas
        for columna, valor in enumerate(header, start=1):
            celda = ws.cell(row=fila_actual, column=columna, value=valor)
            celda.font = Font(bold=True)
        fila_actual += 1

        for fila_datos in datos:
            for columna, valor in enumerate(fila_datos, start=1):
                ws.cell(row=fila_actual, column=columna, value=valor)
            fila_actual += 1

        for columna in range(1, len(header) + 1):
            ws.column_dimensions[ws.cell(row=1, column=columna).column_letter].width = _ANCHO_COLUMNA

    if match_total or match_espesor:
        fila_actual += 1
        if match_total:
            ws.cell(row=fila_actual, column=1, value=match_total.group(0).strip())
            fila_actual += 1
        if match_espesor:
            ws.cell(row=fila_actual, column=1, value=match_espesor.group(0).strip())

    carpeta_destino = "generated"
    os.makedirs(carpeta_destino, exist_ok=True)
    ruta_archivo = os.path.join(carpeta_destino, f"{numero_orden}.xlsx")
    wb.save(ruta_archivo)

    return ruta_archivo
